import openpyxl


class ExcelData:

    # Declaring static in order to use it in a class 'TestEnd2End' without creating an object of the class
    @staticmethod
    def get_excel_dta(test_data):
        # load workbook
        book = openpyxl.load_workbook(r"/E-Commerce-Web-Automation-Testing-Framework/test_data/end2end_data.xlsx")
        # active tab
        sheet = book.active

        # Create a dictionary to store excel data
        data = {}
        data_list = []

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_data:
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_copy = data.copy()
            data_list.append(data_copy)

        return data_list
